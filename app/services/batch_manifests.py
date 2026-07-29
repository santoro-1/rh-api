from __future__ import annotations

import csv
import io
import unicodedata
from dataclasses import dataclass
from datetime import datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class ManifestError(ValueError):
    """The uploaded Excel/CSV manifest cannot be interpreted safely."""


DIGITAL_HUMAN_WORKFLOW = "digital_human"
LTX_LIP_SYNC_WORKFLOW = "ltx_lip_sync"

_HEADER_ALIASES = {
    "row_id": {"row_id", "任务编号", "行编号", "脚本编号"},
    "image_file": {"image_file", "图片文件", "参考图片"},
    "source_video_file": {"source_video_file", "源视频文件", "源视频"},
    "audio_file": {
        "audio_file",
        "音频文件",
        "音频文件（上传模式填写）",
        "总参考音频",
        "总参考音频（上传模式填写）",
        "自定义音频",
    },
    "left_audio_file": {"left_audio_file", "左人物音频", "左边人物音频"},
    "right_audio_file": {"right_audio_file", "右人物音频", "右边人物音频"},
    "prompt": {"prompt", "提示词"},
    "positive_prompt": {
        "positive_prompt",
        "视频正向提示词",
        "画面及对白提示词",
    },
    "speech_script": {
        "speech_script",
        "口播脚本",
        "口播脚本（语音生成模式填写）",
        "口播脚本（上传音频无需填写）",
        "脚本内容",
    },
    "resolution": {"resolution", "最长分辨率"},
    "person_mode": {"person_mode", "单双人模式", "人物模式"},
    "start_time": {"start_time", "开始时间"},
    "end_time": {"end_time", "结束时间"},
    "instance_type": {"instance_type", "运行实例"},
}

_DIGITAL_HEADERS = [
    "row_id",
    "prompt",
]

_LTX_HEADERS = [
    "row_id",
    "speech_script",
]

_SCRIPT_HEADERS = ["row_id", "speech_script"]


@dataclass(frozen=True)
class ParsedManifest:
    workflow_type: str
    rows: list[dict[str, str]]
    source_format: str


def canonical_filename(value: str) -> str:
    """Normalize display filenames only for exact, non-fuzzy lookup."""

    basename = Path(str(value).strip().replace("\\", "/")).name
    return unicodedata.normalize("NFC", basename).casefold()


def template_headers(workflow_type: str) -> list[str]:
    if workflow_type == DIGITAL_HUMAN_WORKFLOW:
        return list(_DIGITAL_HEADERS)
    if workflow_type == LTX_LIP_SYNC_WORKFLOW:
        return list(_LTX_HEADERS)
    raise ManifestError("不支持的批量工作流")


def _normalize_header(value: Any) -> str:
    header = str(value or "").strip()
    folded = header.casefold()
    for canonical, aliases in _HEADER_ALIASES.items():
        if folded in {alias.casefold() for alias in aliases}:
            return canonical
    return ""


def _cell_text(value: Any) -> str:
    if isinstance(value, datetime):
        value = value.time()
    if isinstance(value, time):
        total_minutes = value.hour * 60 + value.minute
        return f"{total_minutes}:{value.second:02d}"
    return str(value or "").strip()


def _rows_from_csv(content: bytes) -> list[list[Any]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ManifestError("CSV 必须使用 UTF-8 编码") from exc
    return [list(row) for row in csv.reader(io.StringIO(text))]


def _rows_from_xlsx(content: bytes) -> list[list[Any]]:
    try:
        workbook = load_workbook(
            io.BytesIO(content),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:  # openpyxl raises several archive/XML exceptions
        raise ManifestError("Excel 文件损坏或不是有效的 .xlsx 文件") from exc
    try:
        sheet = workbook.active
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def parse_manifest(
    filename: str,
    content: bytes,
    workflow_type: str,
    audio_mode: str = "upload",
) -> ParsedManifest:
    """Read one worksheet/CSV into stable internal field names."""

    if workflow_type not in {
        DIGITAL_HUMAN_WORKFLOW,
        LTX_LIP_SYNC_WORKFLOW,
    }:
        raise ManifestError("不支持的批量工作流")

    extension = Path(filename).suffix.lower()
    if extension == ".csv":
        raw_rows = _rows_from_csv(content)
        source_format = "csv"
    elif extension == ".xlsx":
        raw_rows = _rows_from_xlsx(content)
        source_format = "xlsx"
    else:
        raise ManifestError("任务清单只支持 .xlsx 或 .csv")

    nonempty_rows = [
        row for row in raw_rows if any(str(value or "").strip() for value in row)
    ]
    if not nonempty_rows:
        raise ManifestError("任务清单为空")

    normalized_headers = [_normalize_header(value) for value in nonempty_rows[0]]
    seen = [header for header in normalized_headers if header]
    if len(seen) != len(set(seen)):
        raise ManifestError("任务清单包含重复列")

    if audio_mode not in {"upload", "minimax"}:
        raise ManifestError("不支持的音频准备方式")
    if audio_mode == "minimax":
        # Full-flow manifests intentionally contain scripts only. Images or
        # videos are paired by the explicit upload order shown on the page.
        required = set(_SCRIPT_HEADERS)
    else:
        required = (
            {"row_id", "prompt"}
            if workflow_type == DIGITAL_HUMAN_WORKFLOW
            else {"row_id"}
        )
    missing = required - set(normalized_headers)
    if missing:
        raise ManifestError(
            "任务清单缺少必填列：" + "、".join(sorted(missing))
        )
    if (
        audio_mode == "upload"
        and workflow_type == LTX_LIP_SYNC_WORKFLOW
        and not {"speech_script", "positive_prompt"} & set(normalized_headers)
    ):
        raise ManifestError("任务清单缺少必填列：口播脚本")

    rows: list[dict[str, str]] = []
    for source_row_number, raw_row in enumerate(nonempty_rows[1:], start=2):
        values: dict[str, str] = {"source_row_number": str(source_row_number)}
        for index, header in enumerate(normalized_headers):
            if not header:
                continue
            value = raw_row[index] if index < len(raw_row) else ""
            values[header] = _cell_text(value)
        rows.append(values)
    if not rows:
        raise ManifestError("任务清单没有数据行")
    return ParsedManifest(
        workflow_type=workflow_type,
        rows=rows,
        source_format=source_format,
    )


def csv_template(workflow_type: str) -> str:
    output = io.StringIO()
    writer = csv.writer(output, lineterminator="\r\n")
    if workflow_type == "script":
        writer.writerow(["任务编号", "口播脚本"])
        writer.writerow(
            [
                "SCRIPT-001",
                "今天给大家介绍这款产品。请按照一行一条完整脚本填写。",
            ]
        )
    elif workflow_type == DIGITAL_HUMAN_WORKFLOW:
        writer.writerow(["任务编号", "提示词"])
        writer.writerow(
            [
                "TASK-001",
                "人物自然地说话，镜头保持稳定。",
            ]
        )
    elif workflow_type == LTX_LIP_SYNC_WORKFLOW:
        writer.writerow(["任务编号", "口播脚本"])
        writer.writerow(
            [
                "TASK-001",
                "今天给大家介绍这款产品。",
            ]
        )
    else:
        raise ManifestError("不支持的批量工作流")
    return "\ufeff" + output.getvalue()
